# payroll_system/crud/employee_crud.py

from typing import Optional, List
from uuid import UUID
from fastapi import HTTPException
import csv
from io import StringIO
import io
from openpyxl import load_workbook

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError

from models.employee_model import Employee, EmployeeDocument, EmployeeSalaryAssignment
from models.org_models import Organisation  # ✅ Organisation check

from schemas.employee_schemas import (
    EmployeeCreate,
    EmployeeUpdate,
    EmployeeDocumentCreate,
    EmployeeSalaryAssignmentCreate,
)


# ============================================================
# INTERNAL HELPERS
# ============================================================

async def _ensure_organisation_exists(db: AsyncSession, organisation_id: UUID):
    """
    Ensure organisation exists before any employee operation.
    """
    result = await db.execute(
        select(Organisation).where(Organisation.organisation_id == organisation_id)
    )
    org = result.scalar_one_or_none()
    if not org:
        raise ValueError("Organisation not created. Please create organisation first.")


# ============================================================
# EMPLOYEE CRUD
# ============================================================

async def create_employee(db: AsyncSession, emp: EmployeeCreate, organisation_id: UUID):

    await _ensure_organisation_exists(db, organisation_id)

    emp_data = emp.model_dump(exclude_unset=True)
    emp_data["organisation_id"] = organisation_id

    employee = Employee(**emp_data)

    try:
        db.add(employee)
        await db.commit()
        await db.refresh(employee)
        return employee

    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Employee with this email or employee code already exists"
        )

async def get_employee(db: AsyncSession, emp_id: UUID, organisation_id: UUID) -> Optional[Employee]:
    """
    Fetch a single employee by ID within organisation scope.
    Eager-load nested relationships to avoid MissingGreenlet errors.
    """
    await _ensure_organisation_exists(db, organisation_id)

    result = await db.execute(
        select(Employee)
        .where(Employee.employee_id == emp_id, Employee.organisation_id == organisation_id)
        .options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
            selectinload(Employee.location),
            selectinload(Employee.pay_structure),
            selectinload(Employee.manager)
        )
    )
    return result.scalar_one_or_none()


async def list_employees(
    db: AsyncSession,
    organisation_id: UUID,
    department_id: Optional[UUID] = None,
    designation_id: Optional[UUID] = None,
    active_only: bool = True,
) -> List[Employee]:
    """
    List employees for an organisation with optional filters.
    Eager-load relationships.
    """
    await _ensure_organisation_exists(db, organisation_id)

    query = select(Employee).where(Employee.organisation_id == organisation_id).options(
        selectinload(Employee.department),
        selectinload(Employee.designation),
        selectinload(Employee.location),
        selectinload(Employee.pay_structure),
        selectinload(Employee.manager)
    )

    if department_id:
        query = query.where(Employee.department_id == department_id)
    if designation_id:
        query = query.where(Employee.designation_id == designation_id)
    if active_only:
        query = query.where(Employee.is_active.is_(True))

    result = await db.execute(query)
    return result.scalars().all()


async def update_employee(
    db: AsyncSession,
    emp_id: UUID,
    payload: EmployeeUpdate,
    organisation_id: UUID,
) -> Optional[Employee]:
    """
    Update an employee within organisation scope.
    """
    await _ensure_organisation_exists(db, organisation_id)

    employee = await get_employee(db, emp_id, organisation_id)
    if not employee:
        return None

    updates = payload.model_dump(exclude_unset=True)
    for key, value in updates.items():
        setattr(employee, key, value)

    try:
        await db.commit()
        await db.refresh(employee)
        return employee
    except IntegrityError as e:
        await db.rollback()
        raise ValueError("Employee update failed due to integrity error") from e


# ============================================================
# EMPLOYEE DOCUMENTS
# ============================================================

async def upload_employee_document(db: AsyncSession, payload: EmployeeDocumentCreate) -> EmployeeDocument:
    """
    Upload a document for an employee.
    """
    document = EmployeeDocument(**payload.model_dump())

    try:
        db.add(document)
        await db.commit()
        await db.refresh(document)
        return document
    except Exception as e:
        await db.rollback()
        raise ValueError("Employee document upload failed") from e


# ============================================================
# EMPLOYEE SALARY ASSIGNMENTS
# ============================================================

async def assign_salary_structure(db: AsyncSession, payload: EmployeeSalaryAssignmentCreate) -> EmployeeSalaryAssignment:
    """
    Assign a salary structure to an employee.
    Prevent duplicate effective_from entries.
    """
    existing = await db.execute(
        select(EmployeeSalaryAssignment).where(
            EmployeeSalaryAssignment.employee_id == payload.employee_id,
            EmployeeSalaryAssignment.effective_from == payload.effective_from,
        )
    )
    if existing.scalar_one_or_none():
        raise ValueError("Salary assignment with the same effective date already exists")

    assignment = EmployeeSalaryAssignment(**payload.model_dump())

    try:
        db.add(assignment)
        await db.commit()
        await db.refresh(assignment)
        return assignment
    except IntegrityError as e:
        await db.rollback()
        raise ValueError("Failed to assign salary structure") from e


# ============================================================
# EMPLOYEE BULK UPLOAD
# ============================================================

async def bulk_create_employees(db: AsyncSession, file_content, file_type, organisation_id):

    rows = []

    # =========================
    # CSV PARSER
    # =========================
    if file_type == "csv":
        reader = csv.DictReader(io.StringIO(file_content.decode("utf-8")))
        rows = list(reader)

    # =========================
    # EXCEL PARSER
    # =========================
    elif file_type == "xlsx":
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    created = 0
    errors = []

    for idx, row in enumerate(rows, start=1):

        try:
            employee = Employee(
                employee_code=row.get("employee_code"),
                first_name=row.get("first_name"),
                last_name=row.get("last_name"),
                email=row.get("email"),
                phone=row.get("phone"),
                department_id=row.get("department_id"),
                designation_id=row.get("designation_id"),
                organisation_id=organisation_id,
            )

            db.add(employee)
            created += 1

        except Exception as e:
            errors.append({
                "row": idx,
                "error": str(e)
            })

    await db.commit()

    return {
    "total_rows": len(rows),
    "inserted": created,
    "failed": len(errors),
    "errors": errors
}