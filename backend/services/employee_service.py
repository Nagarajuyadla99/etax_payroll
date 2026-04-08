import csv
import io
from uuid import UUID
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import HTTPException

from schemas.employee_schemas import EmployeeCreate
from models.employee_model import Employee

from utils.employee_auth import generate_password
from utils.auth import hash_password
from utils.email_service import send_employee_credentials


# ============================================================
# HELPER FUNCTION
# ============================================================
def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except Exception:
        return None


# ============================================================
# SINGLE EMPLOYEE CREATE (WITH AUTH)
# ============================================================
async def create_employee_with_auth(
    db: AsyncSession,
    emp: EmployeeCreate,
    organisation_id
):
    # ------------------------
    # VALIDATION
    # ------------------------
    if not emp.first_name:
        raise HTTPException(status_code=400, detail="Name is required")

    if not emp.date_of_birth:
        raise HTTPException(status_code=400, detail="Date of birth is required")

    if not emp.email:
        raise HTTPException(status_code=400, detail="Email is required")

    # ------------------------
    # DUPLICATE EMAIL CHECK
    # ------------------------
    result = await db.execute(
        select(Employee).where(
            Employee.email == emp.email,
            Employee.organisation_id == organisation_id
        )
    )

    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email already exists")

    # ------------------------
    # PASSWORD GENERATION
    # ------------------------
    plain_password = generate_password(
        emp.first_name,
        emp.date_of_birth
    )

    hashed_password = hash_password(plain_password)

    # ------------------------
    # CREATE EMPLOYEE
    # ------------------------
    data = emp.model_dump(exclude_unset=True)
    data["organisation_id"] = organisation_id
    data["password_hash"] = hashed_password
    data["is_password_changed"] = False
    employee = Employee(**data)

    db.add(employee)
    await db.commit()
    await db.refresh(employee)

    # ------------------------
    # SEND EMAIL (FAIL SAFE)
    # ------------------------
    try:
        await send_employee_credentials(emp.email, plain_password)
    except Exception as e:
        print("Email sending failed:", str(e))

    return employee


# ============================================================
# BULK CREATE EMPLOYEES
# ============================================================
async def bulk_create_employees(
    db: AsyncSession,
    file_content,
    file_type,
    organisation_id
):

    rows = []

    # ------------------------
    # CSV PARSER
    # ------------------------
    if file_type == "csv":
        reader = csv.DictReader(io.StringIO(file_content.decode("utf-8")))
        rows = list(reader)

    # ------------------------
    # EXCEL PARSER
    # ------------------------
    elif file_type == "xlsx":
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    created = 0
    errors = []

    # ------------------------
    # MAIN LOOP
    # ------------------------
    for idx, row in enumerate(rows, start=1):
        try:
            # ------------------------
            # SAFE TYPE CONVERSIONS
            # ------------------------
            def safe_uuid(val):
                try:
                    return UUID(val) if val else None
                except Exception:
                    return None

            def safe_float(val):
                try:
                    return float(val) if val else None
                except Exception:
                    return None

            def safe_bool(val):
                return str(val).lower() == "true"

            row["department_id"] = safe_uuid(row.get("department_id"))
            row["designation_id"] = safe_uuid(row.get("designation_id"))
            row["location_id"] = safe_uuid(row.get("location_id"))
            row["manager_id"] = safe_uuid(row.get("manager_id"))
            row["pay_structure_id"] = safe_uuid(row.get("pay_structure_id"))

            row["date_of_birth"] = parse_date(row.get("date_of_birth"))
            row["date_of_joining"] = parse_date(row.get("date_of_joining"))
            row["date_of_leaving"] = parse_date(row.get("date_of_leaving"))

            row["annual_ctc"] = safe_float(row.get("annual_ctc"))
            row["is_active"] = safe_bool(row.get("is_active"))

            # ------------------------
            # DUPLICATE CHECK
            # ------------------------
            result = await db.execute(
                select(Employee).where(
                    Employee.organisation_id == organisation_id,
                    Employee.employee_code == row.get("employee_code")
                )
            )

            if result.scalar_one_or_none():
                errors.append({
                    "row": idx,
                    "error": f"Duplicate employee_code: {row.get('employee_code')}"
                })
                continue

            # ------------------------
            # VALIDATION
            # ------------------------
            employee_data = EmployeeCreate(**row)

            data = employee_data.dict(by_alias=True)
            data["organisation_id"] = organisation_id

            # ------------------------
            # PASSWORD GENERATION
            # ------------------------
            if not data.get("first_name") or not data.get("date_of_birth"):
                raise ValueError("Name and DOB required for password")

            plain_password = generate_password(
                data.get("first_name"),
                data.get("date_of_birth")
            )

            hashed_password = hash_password(plain_password)

            data["password_hash"] = hashed_password
            data["is_password_changed"] = False
            # ------------------------
            # CREATE EMPLOYEE
            # ------------------------
            employee = Employee(**data)

            db.add(employee)
            created += 1

        except Exception as e:
            errors.append({
                "row": idx,
                "error": str(e)
            })
            continue

    # ------------------------
    # FINAL COMMIT
    # ------------------------
    await db.commit()

    return {
        "total_rows": len(rows),
        "inserted": created,
        "failed": len(errors),
        "errors": errors
    }